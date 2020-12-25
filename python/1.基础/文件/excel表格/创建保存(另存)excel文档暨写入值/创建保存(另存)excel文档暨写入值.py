#!python3.7
#Filename:创建保存(另存)excel文档暨写入值至单元格.py

import openpyxl

wb = openpyxl.Workbook()    #新建一个空白工作簿（此时未保存下来）
print(wb.sheetnames)    #所有工作表名称列表

sheet = wb.active   #当前工作表
print(sheet.title)
sheet.title = "表1"      #为表命名
print(wb.sheetnames)

#保存新建的工作表，或用于另存excel文件
wb.save("1.xlsx")         

print()

#在工作簿中新建表
class A1:
    wb.create_sheet()       #默认在最后添加,名字为“Sheet*”
    print(wb.sheetnames)
    wb.create_sheet(index = 0,title = "第一个表")   #index创建表格，表格数量从0开始计数
    print(wb.sheetnames)
    wb.create_sheet(index = 2,title = "第三个表")
    print(wb.sheetnames)

    wb.save("2.xlsx")    #excel另存时，名字要改变，否则会直接在原表格进行更改

print()

#在工作簿中删除表
class A2:
    print(wb.sheetnames)
    wb.remove(wb["Sheet"])      #wb.remove_sheet(wb["Sheet"])
    del wb["表1"]               #wb.remove_sheet(wb["表1"])
    print(wb.sheetnames)
    
    wb.save("3.xlsx")

print()

#将值写入单元格
class A3:        #类似于将value写入字典
    #i = wb["表1"]        #每次excel保存（或另存）后，wb都会默认指向最新存储的excel
    i = wb["第一个表"]
    i["B2"] = "Hello"
    print(i["B2"].value)
    wb.save("4.xlsx")
    
    i = wb["第三个表"]
    i["B2"] = "World"
    print(i["B2"].value)
    wb.save("4.xlsx")


print("\nDone")

"""
当需要打开已有的excel时，使用
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["表名"]
不要忘记保存
"""



