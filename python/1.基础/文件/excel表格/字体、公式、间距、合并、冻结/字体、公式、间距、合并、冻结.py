#!python3.7
#Filename:字体、公式、间距、合并、冻结.py

#注：要运行哪个函数，在末尾进行更改数字即可

import openpyxl

wb = openpyxl.Workbook()
sheet = wb["Sheet"]         #新建的excel只有一个表格，默认名称为“Sheet”
sheet.title = "表1"
for i in range(1,10):
    sheet["A"+str(i)] = 0
    sheet["B"+str(i)] = 0
    sheet["C"+str(i)] = 0
wb.save("example.xlsx")


#字体
def A1():
    """
字体
    """
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb["表1"]
    italic24Font = Font(size=24,italic=True)
    sheet["B2"].font = italic24Font             #注意大小写

    sheet["B2"] = "World"
    sheet["A1"] = "Hello"
    wb.save("字体.xlsx")
    """
name——字符串——字体名称，如“Calibri”或“黑体”
size——整型——字的大小
bold——布尔型——True，粗体字体
italic——布尔型——True，斜体字体
    """

#excel内公式
def A2():
    """
excel内公式
    """
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb["表1"]
    for i in range(1,10):
        sheet["C"+str(i)] = i           #字符型只能和字符型相加组合
    sheet["C10"] = "=SUM(C1:C9)"    #反映到表格内为一个公式（存储为C1到C9的和）
    wb.save("excel内公式.xlsx")
    for i_row in range(1,11):
        print("C"+str(i_row),"：",sheet.cell(row=i_row,column=3).value)


#间距
def A3():
    """
间距
    """
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb["表1"]
    sheet.row_dimensions[1].height = 70         #设置第1行的高度（0到409间的浮点数，表格初始默认12.75
    sheet.column_dimensions["B"].width = 20     #设置B列的宽度（0到255间的浮点数，表格初始默认8.43
    wb.save("行间距.xlsx")


#(取消)合并单元格
def A4():
    """
(取消)合并单元格
    """
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb["表1"]
    sheet.merge_cells("A1:D3")          #合并单元格
    sheet.merge_cells("C5:D5")
    sheet["A1"] = "十二个单元格合并"    #为合并的单元格设置值时，只需设置合并组的左上角的单元格
    sheet["C5"] = "两个单元格合并"
    wb.save("合并单元格.xlsx")

    wb = openpyxl.load_workbook("合并单元格.xlsx")
    sheet = wb["表1"]
    sheet.unmerge_cells("A1:D3")        #取消合并的单元格
    wb.save("取消合并的单元格.xlsx")


#冻结窗格
def A5():
    """
冻结窗格
    """
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb["表1"]
    sheet.freeze_panes = "B2"       #冻结A列和第一行
    #sheet.freeze_panes = "B1"       #冻结A列（会覆盖上一次的冻结操作）
    sheet.freeze_panes = "C4"       #冻结A、B列和一、二、三行
    #sheet.freeze_panes = "A1"       #相当于解除所有冻结（也可sheet.freeze_pans = None）
    wb.save("冻结窗格.xlsx")
#会冻结指定单元格的左列和上行，但不包含本身所在的行列

if __name__ == "__main__":
    A5()
    print("运行了：",A5.__doc__)    #函数的文档字符串
   
   
print("\nDone")
