#!python3.7
#Filename:excel表格功能预览.py
#第三方库：openpyxl

import openpyxl

#读取指定的工作簿
wb = openpyxl.load_workbook("example.xlsx")     
#print(type(wb))

"""
print(wb.get_sheet_names())         #此函数已弃用
print(wb.sheetnames)            #获取工作簿中所有工作表名称的列表
"""

#获取工作簿中的活动工作表（位于工作簿当前页的工作表）
sheet = wb.active       
print(sheet)
"""
#print(wb.get_sheet_by_name("Sheet3"))   #此函数已弃用
somesheet = wb["3"]      #通过将工作表名称字符串传递来获取该对象
print(somesheet)
print(somesheet.title)      #返回工作表的名称
"""

"""
print()
class A1:
    #value返回指定工作表内指定单元格的内容
    print(sheet["A1"].value)   
    b1 = sheet["B1"]
    print(b1.value)
    print(sheet["C1"].value)

    prt = "单元格" + b1.coordinate + "是" + b1.value     #coordinate属性提供了“B1”
    #prt1 = str(b1.row) + "行，" + b1.column + "列是" + b1.value     #row提供了“1”，column提供了“B”
    print(prt)
    #print(prt1)

    #cell方法获取指定单元格,row指定行，column指定列
    ce = sheet.cell(row=1,column=2)     #为row,column传递关键字参数（大于0的整数）
    print(ce)
    print(ce.value)
    for i in range(1,8,2):
        print(i,sheet.cell(row=i,column=2).value)

    print(sheet.max_row,sheet.max_column)   #确定工作表的大小（多少行，多少列）

"""

"""
print()
class A2:
    #在列的字母和数字之间转换
    from openpyxl.cell import get_column_letter,column_index_from_string
    print(get_column_letter(1))     #将列数字转换为字母
    print(get_column_letter(27))
    print(get_column_letter(sheet.max_column))

    print(column_index_from_string("AB"))   #将列字母转换为数字
"""

print()
class A3:
    #从表格获取行和列
    tu = sheet["A1":"C3"]   #返回一个元组，内每个元素为一个行所有单元格组成的元组，即有几行则有几个元素
    print(tu)

    for rowcell in sheet["A1":"C3"]:    #rowcell指代每行
        for rows in rowcell:     #此时row指代了单元格
            print(rows.coordinate,rows.value)
        print("——————")

    co = sheet.columns[1]   #列columns从0开始（行raws从1开始），可返回特定对象元组
    print(co)
    for i in co:
        print(i.value)






















    
